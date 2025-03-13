import telebot
from config import token
from telebot import types
from aiohttp import web 
import ssl 
import paramiko                    
from paramiko import SSHClient    
from paramiko import AutoAddPolicy 
import time
import datetime
from datetime import datetime
from redminelib import Redmine
from datetime import datetime
from datetime import timedelta
import wget
import glob
import os.path
import os
import shutil
import xlwings
import openpyxl
import keyboard.all as kball
import keyboard.mvu as kbmvu
import keyboard.ru as kbru
import keyboard.du as kbdu
import keyboard.tu as kbtu
import keyboard.svu as kbsvu
import keyboard.sku as kbsku
import logger as log
import pymysql
mydb = pymysql.connect(host="input_ip_mysql", user="input_login_mysql", passwd="input_pass_mysql", database="input_name_mysql", port=int(input_port_mysql), cursorclass=pymysql.cursors.DictCursor)

localfolder = "" 
srczayavka = "None"
redmineoffice = Redmine('input_url_redmine', username='input_login_redmine', password='input_pass_redmine')
time = datetime.now()
checktime7 = time -timedelta(days = 7)
checktime14 = time -timedelta(days = 14)
checktime30 = time -timedelta(days = 30)

WEBHOOK_LISTEN = "0.0.0.0"
WEBHOOK_PORT = input_port_webhook

WEBHOOK_SSL_CERT = "fullchain.pem"
WEBHOOK_SSL_PRIV = "privkey.pem"

bot = telebot.TeleBot(token)
bot_username = bot.get_me().username
app = web.Application()
async def handle(request):
    if request.match_info.get("token") == bot.token:
        request_body_dict = await request.json()
        update = telebot.types.Update.de_json(request_body_dict)
        bot.process_new_updates([update])
        return web.Response()
    else:
        return web.Response(status=403)

app.router.add_post("/{token}/", handle)

user_dict = {}
class User:
    def __init__(self, name):
        self.name = name
        self.typekassir = None
        self.phone = None
        self.uchastok = None
        self.city = None
        self.kkt = None
        self.idkkt = None
        self.situation = None
        self.confirm = None
bankuser_dict = {}
class User:
    def __init__(self, name):
        self.name = name
        self.banktypequestion = None
        self.bankconfirm = None
        self.bankanon = None
bkfoto_dict = {}
class User:
    def __init__(self, name):
        self.name = name
        self.bkfoto_station = None
        self.bkfoto_type = None
        self.bkfoto_confirm = None
rbkfotopl_dict = {}
class User:
    def __init__(self, name):
        self.name = name
        self.rbkfotopl_confirm = None
rbkfotost_dict = {}
class User:
    def __init__(self, name):
        self.name = name
        self.rbkfotost_confirm = None
rbkfotoother_dict = {}
class User:
    def __init__(self, name):
        self.name = name
        self.rbkfotoother_confirm = None
rbkfotokdo_dict = {}
class User:
    def __init__(self, name):
        self.name = name
        self.rbkfoto_station = None
        self.rbkfotokdo_confirm = None
ekspfoto_dict = {}
class User:
    def __init__(self, name):
        self.name = name
        self.ekspfoto_type = None
        self.ekspfoto_confirm = None
redmine_dict = {}
class User:
    def __init__(self, name):
        self.name = name
        self.redminepass = None
        self.redmineuser = None
abon_dict = {}
class User:
    def __init__(self, name):
        self.name = name
        self.abon_num = None

@bot.message_handler(commands=['start'])
def start(message: types.Message):

    import datetime
    chat_id = message.chat.id
    firstname = message.from_user.first_name 
    username = message.from_user.username
    if username is None:
        username = "null"
    if firstname is None:
        firstname = "null"
    mycursor = mydb.cursor() 
    mydb.ping(reconnect=True) 
    mycursor.execute("SELECT area FROM users WHERE chat_id=%s", (chat_id))
    data = "error"
    for i in mycursor:
        data = i 
    if data == "error": 
        bot.send_message(chat_id, '🐙 Привет ' + firstname + '!\nДля получения доступа перешли свой ID начальнику своего участка', reply_markup=types.ReplyKeyboardRemove())
        bot.send_message(chat_id, str(chat_id), reply_markup=types.ReplyKeyboardRemove())
        log.writelog(user_id=message.chat.id, username=message.from_user.first_name, text="/start Запрос на первичную активацию >> ")
        bot.send_message(nmvu, '⚠️Попытка первичной активации!\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username + '\n🔄 Отправлено предложение о запросе на активацию.')
        mydb.close()
    else:
        mycursor = mydb.cursor()  
        mycursor.execute("SELECT * FROM users WHERE chat_id=%s", (chat_id))
        result = mycursor.fetchall()
        for row in result:
            valid_area = row['area']
            valid_comment = row['comment']
        mydb.commit()
        if valid_area == "del":
            bot.send_message(chat_id, '🐙 Привет ' + firstname + '!\nДоступ был закрыт!\n' + str(chat_id), reply_markup=types.ReplyKeyboardRemove())
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text=" /start БЛОКИРОВКА. ПОЛЬЗОВАТЕЛЬ БЫЛ УДАЛЁН >>") 
            mydb.close()
        elif valid_area == "mvu":
            dtn = datetime.datetime.now()
            time = dtn.strftime("%d-%m-%Y %H:%M") 
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'" 
            mycursor.execute(nick)
            mycursor.execute(add_last_use) 
            mycursor.execute(add_firstname)
            mydb.commit()
            mydb.close()
            bot.send_sticker(message.chat.id, 'CAACAgIAAxkBAAEDcTlhsbT2Pf_KfjLLo5ssWosfFsI6KAACngIAAzigCnbbnChT4sv5IwQ', reply_markup=types.ReplyKeyboardRemove())
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text=" /start активация МВУ >>")
            bot.send_message(nmvu, text=f'🟢 Активация МВУ start \nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        elif valid_area == "ru":
            dtn = datetime.datetime.now()
            time = dtn.strftime("%d-%m-%Y %H:%M") 
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'" 
            mycursor.execute(nick)
            mycursor.execute(add_last_use)
            mycursor.execute(add_firstname)
            mydb.commit()
            mydb.close()
            bot.send_sticker(message.chat.id, 'CAACAgIAAxkBAAEDcTlhsbT2Pf_KfjLLo5ssWosfFsI6KAACngIAAzigCnbbnChT4sv5IwQ', reply_markup=types.ReplyKeyboardRemove())
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text=" /start активация РУ >>")  
            bot.send_message(nru, text=f'🟢 Активация РУ start \nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        elif valid_area == "du":
            dtn = datetime.datetime.now() 
            time = dtn.strftime("%d-%m-%Y %H:%M") 
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'" 
            mycursor.execute(nick) 
            mycursor.execute(add_last_use) 
            mycursor.execute(add_firstname)  
            mydb.commit()
            mydb.close()
            bot.send_sticker(message.chat.id, 'CAACAgIAAxkBAAEDcTlhsbT2Pf_KfjLLo5ssWosfFsI6KAACngIAAzigCnbbnChT4sv5IwQ', reply_markup=types.ReplyKeyboardRemove())
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text=" /start активация  ДУ>>")  
            bot.send_message(ndu, text=f'🟢 Активация ДУ start \nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        elif valid_area == "tu":
            dtn = datetime.datetime.now() 
            time = dtn.strftime("%d-%m-%Y %H:%M") 
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'" 
            mycursor.execute(nick) 
            mycursor.execute(add_last_use) 
            mycursor.execute(add_firstname)  
            mydb.commit()
            mydb.close()
            bot.send_sticker(message.chat.id, 'CAACAgIAAxkBAAEDcTlhsbT2Pf_KfjLLo5ssWosfFsI6KAACngIAAzigCnbbnChT4sv5IwQ', reply_markup=types.ReplyKeyboardRemove())
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text=" /start активация  ТУ>>") 
            bot.send_message(ntu, text=f'🟢 Активация ТУ start \nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        elif valid_area == "svu":
            dtn = datetime.datetime.now() 
            time = dtn.strftime("%d-%m-%Y %H:%M") 
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'"
            mycursor.execute(nick) 
            mycursor.execute(add_last_use)
            mycursor.execute(add_firstname) 
            mydb.commit()
            mydb.close()
            bot.send_sticker(message.chat.id, 'CAACAgIAAxkBAAEDcTlhsbT2Pf_KfjLLo5ssWosfFsI6KAACngIAAzigCnbbnChT4sv5IwQ', reply_markup=types.ReplyKeyboardRemove())
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text=" /start активация  СВУ>>")  
            bot.send_message(nsvu, text=f'🟢 Активация СВУ start \nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        elif valid_area == "sku":
            dtn = datetime.datetime.now()
            time = dtn.strftime("%d-%m-%Y %H:%M")
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'"
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'"
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'" 
            mycursor.execute(nick) 
            mycursor.execute(add_last_use) 
            mycursor.execute(add_firstname) 
            mydb.commit()
            mydb.close()
            bot.send_sticker(message.chat.id, 'CAACAgIAAxkBAAEDcTlhsbT2Pf_KfjLLo5ssWosfFsI6KAACngIAAzigCnbbnChT4sv5IwQ', reply_markup=types.ReplyKeyboardRemove())
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text=" /start активация  СКУ>>")  
            bot.send_message(nsku, text=f'🟢 Активация СКУ start \nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        else:
            bot.send_message(chat_id, '🐙 Привет ' + firstname + '!\nДля получения доступа перешли свой ID начальнику своего участка', reply_markup=types.ReplyKeyboardRemove())
            bot.send_message(chat_id, str(chat_id), reply_markup=types.ReplyKeyboardRemove())
            log.writelog(user_id=message.chat.id, username=message.from_user.first_name, text=" /start У ПОЛЬЗОВАТЕЛЯ НЕТ ПЕРЕМЕННОЙ В AREA >>")  
            bot.send_message(nmvu,'⛔Попытка активации start после блокировки!\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username + '\n🔄 Отправлено предложение о запросе на активацию.')
            mydb.close()

@bot.message_handler(commands=['usr'])
def usr(message: types.Message):
    import datetime
    chat_id = message.chat.id 
    firstname = message.from_user.first_name 
    username = message.from_user.username  
    if username is None:
        username = "null"
    if firstname is None:
        firstname = "null"
    mycursor = mydb.cursor() 
    mydb.ping(reconnect=True)
    mycursor.execute("SELECT area FROM users WHERE chat_id=%s", (chat_id)) 
    data = "error"  
    for i in mycursor:
        data = i  
    if data == "error":
      
        log.writelog(user_id=message.chat.id, username=message.from_user.first_name, text="/usr БЛОКИРОВКА ПОЛЬЗОВАТЕЛЯ НЕТ В БАЗЕ >> ")  
        bot.send_message(nmvu, '⛔ Попытка новой активации 👤 usr!\n⚠️Пользователя нет в базе\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        mydb.close()
    else:
        mycursor = mydb.cursor() 
        mycursor.execute("SELECT * FROM users WHERE chat_id=%s", (chat_id))
        result = mycursor.fetchall()
        for row in result:
            valid_area = row['area']
            valid_comment = row['comment']
            valid_usr = row['usr']
        mydb.commit()
        if valid_area == "del":
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/usr БЛОКИРОВКА. ПОЛЬЗОВАТЕЛЬ БЫЛ УДАЛЁН >>") 
            bot.send_message(nmvu,'⛔Попытка активации 👤 usr\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username + '\nПользователь был удалён')
            mydb.close()
        elif valid_usr == "0":
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/usr БЛОКИРОВКА. ПОЛЬЗОВАТЕЛЬ БЫЛ УДАЛЁН >>")  
            bot.send_message(nmvu, '⛔Попытка активации 👤 usr\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username + '\nПользователь был удалён')
            mydb.close()
        elif(valid_area == "mvu" and valid_usr == 1):
            dtn = datetime.datetime.now() 
            time = dtn.strftime("%d-%m-%Y %H:%M") 
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'" 
            mycursor.execute(nick) 
            mycursor.execute(add_last_use)
            mycursor.execute(add_firstname) 
            mydb.commit()
            mydb.close()
            bot.send_message(chat_id,'🐙 Меню МВУ:',parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
            bot.send_message(chat_id, 'работа с пользователями.💫', parse_mode="Markdown", reply_markup=kball.mmusr)
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/usr МВУ активация >>")
            bot.send_message(nmvu, text=f'🟣 Активация 👤 usr МВУ\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        elif(valid_area == "ru" and valid_usr == 1):
            dtn = datetime.datetime.now() 
            time = dtn.strftime("%d-%m-%Y %H:%M") 
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'"
            mycursor.execute(nick) 
            mycursor.execute(add_last_use)
            mycursor.execute(add_firstname)
            mydb.commit()
            mydb.close()
            bot.send_message(chat_id,'🐙 Меню РУ:',parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
            bot.send_message(chat_id, 'работа с пользователями.💫', parse_mode="Markdown", reply_markup=kball.mmusr)
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/usr РУ активация >>") 
            bot.send_message(nru, text=f'🟣 Активация 👤 usr РУ\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        elif(valid_area == "du" and valid_usr == 1):
            dtn = datetime.datetime.now()
            time = dtn.strftime("%d-%m-%Y %H:%M")
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'"
            mycursor.execute(nick)
            mycursor.execute(add_last_use)
            mycursor.execute(add_firstname)
            mydb.commit()
            mydb.close()
            bot.send_message(chat_id,'🐙 Меню ДУ:',parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
            bot.send_message(chat_id, 'работа с пользователями.💫', parse_mode="Markdown", reply_markup=kball.mmusr)
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/usr ДУ активация >>")
            bot.send_message(du, text=f'🟣 Активация 👤 usr ДУ\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        elif(valid_area == "tu" and valid_usr == 1):
            dtn = datetime.datetime.now()
            time = dtn.strftime("%d-%m-%Y %H:%M")
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'"
            mycursor.execute(nick)
            mycursor.execute(add_last_use)
            mycursor.execute(add_firstname)
            mydb.commit()
            mydb.close()
            bot.send_message(chat_id,'🐙 Меню ТУ:',parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
            bot.send_message(chat_id, 'работа с пользователями.💫', parse_mode="Markdown", reply_markup=kball.mmusr)
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/usr ТУ активация >>")
            bot.send_message(tu, text=f'🟣 Активация 👤 usr ТУ\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        elif(valid_area == "svu" and valid_usr == 1):
            dtn = datetime.datetime.now()
            time = dtn.strftime("%d-%m-%Y %H:%M")
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'"
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'"
            mycursor.execute(nick)
            mycursor.execute(add_last_use)
            mycursor.execute(add_firstname)
            mydb.commit()
            mydb.close()
            bot.send_message(chat_id,'🐙 Меню СВУ:',parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
            bot.send_message(chat_id, 'работа с пользователями.💫', parse_mode="Markdown", reply_markup=kball.mmusr)
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/usr СВУ активация >>")
            bot.send_message(nsvu, text=f'🟣 Активация 👤 usr СВУ\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        elif(valid_area == "sku" and valid_usr == 1):
            dtn = datetime.datetime.now()
            time = dtn.strftime("%d-%m-%Y %H:%M") 
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'" 
            mycursor.execute(nick) 
            mycursor.execute(add_last_use) 
            mycursor.execute(add_firstname) 
            mydb.commit()
            mydb.close()
            bot.send_message(chat_id,'🐙 Меню СКУ:',parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
            bot.send_message(chat_id, 'работа с пользователями.💫', parse_mode="Markdown", reply_markup=kball.mmusr)
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/usr СКУ активация >>") 
            bot.send_message(nsku, text=f'🟣 Активация 👤 usr СКУ\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        else:
            bot.send_message(chat_id, '🐙 Привет ' + firstname + '!\nТебе сюда нельзя!\nДля получения доступа перешли свой ID начальнику своего участка', reply_markup=types.ReplyKeyboardRemove())
            bot.send_message(chat_id, str(chat_id), reply_markup=types.ReplyKeyboardRemove())
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/usr ДОСТУП ЗАБЛОКИРОВАН НЕТ ПРАВ>>")  
            bot.send_message(nmvu,'⛔Попытка активации 👤 usr\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username + '\n🔄 Отправлено предложение о запросе на активацию.')

@bot.message_handler(commands=['adm'])
def adm(message: types.Message):
    chat_id = message.chat.id
    firstname = message.from_user.first_name 
    username = message.from_user.username 
    if username is None:
        username = "null"
    if firstname is None:
        firstname = "null"
    mycursor = mydb.cursor()  
    mydb.ping(reconnect=True) 
    mycursor.execute("SELECT adm FROM users WHERE chat_id=%s", (chat_id))
    data = "error"  
    for i in mycursor:
        data = i 
    if data == "error": 
        bot.send_message(admin, '⛔ Попытка новой активации 🐙 adm!\n⚠️Пользователя нет в базе\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        log.writelog(user_id=message.chat.id, username=message.from_user.first_name, text=" /adm блокировка >>") 
        mydb.close()
    else:
        mycursor.execute("SELECT * FROM users WHERE chat_id=%s", (chat_id))
        result = mycursor.fetchall()
        for row in result:
            valid_user = row['adm']
            valid_comment = row['comment']
        if valid_user == 1:
            import datetime
            dtn = datetime.datetime.now() 
            time = dtn.strftime("%d-%m-%Y %H:%M")
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'" 
            firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'" 
            mycursor.execute(nick) 
            mycursor.execute(last_use)
            mycursor.execute(firstname) 
            mydb.commit()
            mydb.close()
            bot.send_sticker(message.chat.id, 'CAACAgIAAxkBAAEDf1FhuvYrjvw2kGP2JoJDpVIu-T6N0wACwAADVp29Ct1dnTI9q-YvIwQ')
            bot.send_message(chat_id,'🐙 Меню административное.💫',parse_mode="Markdown", reply_markup=kball.mmadm)
            log.writelog(user_id=message.chat.id, username=valid_comment, text=" /adm активация >>")
        else:
            bot.send_message(admin, '⛔ Блокировка активации 🐙 adm!\n⚠️Доступ запрещён\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
            log.writelog(user_id=message.chat.id, username=message.from_user.first_name, text=" /adm блокировка >>")
            mydb.close()

@bot.message_handler(commands=['cam'])
def cam(message: types.Message):
    import datetime 
    chat_id = message.chat.id
    firstname = message.from_user.first_name
    username = message.from_user.username 
    if username is None:
        username = "null"
    if firstname is None:
        firstname = "null"
    mycursor = mydb.cursor() 
    mydb.ping(reconnect=True)  
    mycursor.execute("SELECT cam FROM users WHERE chat_id=%s", (chat_id)) 
    data = "error"  
    for i in mycursor:
        data = i 
    if data == "error": 
        bot.send_message(admin, '⛔ Попытка новой активации 📸 cam!\n⚠️Пользователя нет в базе\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        bot.send_message(nmvu, '⛔ Попытка новой активации 📸 cam!\n⚠️Пользователя нет в базе\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        log.writelog(user_id=message.chat.id, username=message.from_user.first_name, text=" /cam НЕТ В БАЗЕ >>") 
        mydb.close()
    else:
        mycursor.execute("SELECT * FROM users WHERE chat_id=%s", (chat_id))
        result = mycursor.fetchall()
        for row in result:
            valid_user = row['cam']
            valid_comment = row['comment']
        if valid_user == 1:
            dtn = datetime.datetime.now() 
            time = dtn.strftime("%d-%m-%Y %H:%M") 
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'" 
            firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'" 
            mycursor.execute(nick)
            mycursor.execute(last_use)
            mycursor.execute(firstname)
            mydb.commit()
            mydb.close()
            bot.send_message(chat_id,'🐙 Меню камер.💫',parse_mode="Markdown", reply_markup=kball.mmcam)
            log.writelog(user_id=message.chat.id, username=message.from_user.first_name, text=" /cam активация >>") 
            bot.send_message(admin, text=f'🔴 Активация 📸 cam \nИмя: '+ valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
            bot.send_message(nmvu, text=f'🔴 Активация 📸 cam \nИмя: '+ valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        else:
            bot.send_message(admin, '⛔ Блокировка активации 📸 cam!\n⚠️Доступ запрещён\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
            bot.send_message(nmvu, '⛔ Блокировка активации 📸 cam!\n⚠️Доступ запрещён\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
            log.writelog(user_id=message.chat.id, username=message.from_user.first_name, text=" /cam НЕТ ДОСТУПА >>")
            mydb.close()

@bot.message_handler(commands=['send'])
def send(message: types.Message):
    import datetime 
    chat_id = message.chat.id
    firstname = message.from_user.first_name
    username = message.from_user.username 
    if username is None:
        username = "null"
    if firstname is None:
        firstname = "null"
    mycursor = mydb.cursor() 
    mydb.ping(reconnect=True) 
    mycursor.execute("SELECT area FROM users WHERE chat_id=%s", (chat_id))
    data = "error"  
    for i in mycursor:
        data = i
    if data == "error": 
        log.writelog(user_id=message.chat.id, username=message.from_user.first_name, text="/send БЛОКИРОВКА ПОЛЬЗОВАТЕЛЯ НЕТ В БАЗЕ >> ")  
        bot.send_message(admin, '⛔ Попытка новой активации ✉ send!\n⚠️Пользователя нет в базе\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        bot.send_message(nmvu, '⛔ Попытка новой активации ✉ send!\n⚠️Пользователя нет в базе\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        mydb.close()
    else:
        mycursor = mydb.cursor() 
        mycursor.execute("SELECT * FROM users WHERE chat_id=%s", (chat_id))
        result = mycursor.fetchall()
        for row in result:
            valid_area = row['area']
            valid_comment = row['comment']
            valid_send = row['send']
        mydb.commit()
        if valid_area == "del":
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/send БЛОКИРОВКА. ПОЛЬЗОВАТЕЛЬ БЫЛ УДАЛЁН >>") 
            bot.send_message(admin,'⛔️Попытка активации ✉ send!\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username + '\nПользователь был удалён')
            bot.send_message(nmvu,'⛔Попытка активации ✉ send!\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username + '\nПользователь был удалён')
            mydb.close()
        elif valid_send == "0":
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/send БЛОКИРОВКА. ПОЛЬЗОВАТЕЛЬ БЫЛ УДАЛЁН >>")  
            bot.send_message(admin, '⛔️Попытка активации ✉ send!\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username + '\nПользователь был удалён')
            bot.send_message(nmvu, '⛔Попытка активации ✉ send!\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username + '\nПользователь был удалён')
            mydb.close()
        elif(valid_area == "mvu" and valid_send == 1):
            dtn = datetime.datetime.now() 
            time = dtn.strftime("%d-%m-%Y %H:%M") 
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'" 
            mycursor.execute(nick) 
            mycursor.execute(add_last_use) 
            mycursor.execute(add_firstname)  
            mydb.commit()
            mydb.close()
            bot.send_message(chat_id,'🐙 Меню МВУ:',parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
            bot.send_message(chat_id, 'работа с рассылкой.💫', parse_mode="Markdown", reply_markup=kball.areasend)
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/send МВУ активация >>")  
            bot.send_message(admin, text=f'🔵 Активация ✉ send!\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
            bot.send_message(nmvu, text=f'🔵 Активация ✉ send!\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        elif(valid_area == "ru" and valid_send == 1):
            dtn = datetime.datetime.now() 
            time = dtn.strftime("%d-%m-%Y %H:%M") 
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'" 
            mycursor.execute(nick) 
            mycursor.execute(add_last_use) 
            mycursor.execute(add_firstname)  
            mydb.commit()
            mydb.close()
            bot.send_message(chat_id,'🐙 Меню РУ:',parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
            bot.send_message(chat_id, 'работа с рассылкой.💫', parse_mode="Markdown", reply_markup=kball.areasend)
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/send РУ активация >>")  
            bot.send_message(admin, text=f'🔵 Активация ✉ send!\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
            bot.send_message(nru, text=f'🔵 Активация ✉ send!\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        elif(valid_area == "du" and valid_send == 1):
            dtn = datetime.datetime.now()
            time = dtn.strftime("%d-%m-%Y %H:%M") 
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'" 
            mycursor.execute(nick) 
            mycursor.execute(add_last_use) 
            mycursor.execute(add_firstname) 
            mydb.commit()
            mydb.close()
            bot.send_message(chat_id,'🐙 Меню ДУ:',parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
            bot.send_message(chat_id, 'работа с рассылкой.💫', parse_mode="Markdown", reply_markup=kball.areasend)
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/send ДУ активация >>") 
            bot.send_message(admin, text=f'🔵 Активация ✉ send!\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
            bot.send_message(ndu, text=f'🔵 Активация ✉ send!\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        elif(valid_area == "tu" and valid_send == 1):
            dtn = datetime.datetime.now()
            time = dtn.strftime("%d-%m-%Y %H:%M") 
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'"
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'"
            mycursor.execute(nick) 
            mycursor.execute(add_last_use) 
            mycursor.execute(add_firstname)  
            mydb.commit()
            mydb.close()
            bot.send_message(chat_id,'🐙 Меню ТУ:',parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
            bot.send_message(chat_id, 'работа с рассылкой.💫', parse_mode="Markdown", reply_markup=kball.areasend)
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/send ТУ активация >>")  
            bot.send_message(admin, text=f'🔵 Активация ✉ send!\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
            bot.send_message(ntu, text=f'🔵 Активация ✉ send!\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        elif(valid_area == "svu" and valid_send == 1):
            dtn = datetime.datetime.now() 
            time = dtn.strftime("%d-%m-%Y %H:%M") 
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'" 
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'"
            mycursor.execute(nick)
            mycursor.execute(add_last_use)
            mycursor.execute(add_firstname)
            mydb.commit()
            mydb.close()
            bot.send_message(chat_id,'🐙 Меню СВУ:',parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
            bot.send_message(chat_id, 'работа с рассылкой.💫', parse_mode="Markdown", reply_markup=kball.areasend)
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/send СВУ активация >>")
            bot.send_message(admin, text=f'🔵 Активация ✉ send!\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
            bot.send_message(nsvu, text=f'🔵 Активация ✉ send!\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        elif(valid_area == "sku" and valid_send == 1):
            dtn = datetime.datetime.now()
            time = dtn.strftime("%d-%m-%Y %H:%M") 
            nick = "UPDATE users set nick = '" + username + "' WHERE chat_id = '" + str(chat_id) + "'"
            add_last_use = "UPDATE users set last_use = '" + time + "' WHERE chat_id = '" + str(chat_id) + "'"
            add_firstname = "UPDATE users set firstname = '" + firstname + "' WHERE chat_id = '" + str(chat_id) + "'"
            mycursor.execute(nick) 
            mycursor.execute(add_last_use) 
            mycursor.execute(add_firstname) 
            mydb.commit()
            mydb.close()
            bot.send_message(chat_id,'🐙 Меню СКУ:',parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
            bot.send_message(chat_id, 'работа с рассылкой.💫', parse_mode="Markdown", reply_markup=kball.areasend)
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/send СКУ активация >>") 
            bot.send_message(admin, text=f'🔵 Активация ✉ send!\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
            bot.send_message(nsku, text=f'🔵 Активация ✉ send!\nИмя: ' + valid_comment + '\nID: ' + str(message.chat.id) + "\nНик: @" + username)
        else:
            bot.send_message(chat_id, '🐙 Привет ' + firstname + '!\nТебе сюда нельзя!\nДля получения доступа перешли свой ID начальнику своего участка', reply_markup=types.ReplyKeyboardRemove())
            bot.send_message(chat_id, str(chat_id), reply_markup=types.ReplyKeyboardRemove())
            log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id, username=message.from_user.first_name, text="/send ДОСТУП ЗАБЛОКИРОВАН НЕТ ПРАВ>>") 
            bot.send_message(admin,'⛔️Попытка активации 👤 usr\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username + '\n🔄 Отправлено предложение о запросе на активацию.')
            bot.send_message(nmvu,'⛔Попытка активации 👤 usr\nИмя: ' + message.from_user.first_name + '\nID: ' + str(message.chat.id) + "\nНик: @" + username + '\n🔄 Отправлено предложение о запросе на активацию.')
            mydb.close()

@bot.callback_query_handler(func=lambda call: True)
def handler_call(call):
    import datetime
    import time 
    global chat_id
    import pymysql 
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    mydb.ping(reconnect=True)
    mycursor = mydb.cursor() 
    mycursor.execute("SELECT area FROM users WHERE chat_id=%s", (chat_id))  
    data = "error" 
    for i in mycursor:
        data = i 
    if data == "error":  
        log.writelog(user_id=call.message.chat.id, username=call.from_user.first_name, text="БЛОКИРОВКА! ПОЛЬЗОВАТЕЛЯ НЕТ В БАЗЕ >> ") 
        mydb.close()
    else:
        mycursor = mydb.cursor()  
        mycursor.execute("SELECT * FROM users WHERE chat_id=%s", (chat_id))
        result = mycursor.fetchall()
        for row in result:
            valid_area = row['area']
            valid_comment = row['comment']
        mydb.commit()
        mydb.close()
        if valid_area == "del":
            log.writelog_start(valid_comment=valid_comment, user_id=call.message.chat.id, username=call.from_user.first_name, text="БЛОКИРОВКА! ПОЛЬЗОВАТЕЛЬ БЫЛ УДАЛЁН >>") 
        elif valid_area in ["mvu", "ru", "du", "tu", "svu", "sku"]:
            if call.data == 'admusr_opros':
                
                bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                      text='📊 Меню работы с опросами.\nУчасток:', parse_mode='Markdown',
                                      reply_markup=kball.admopros)
            if call.data == 'admopros':
                msg = bot.send_message(chat_id=chat_id,
                                       text='🐙 Показать настройки опроса?\n\n❌ стоп - для отмены\n✅ далее - для продолжения',
                                       reply_markup=kball.stopandnext)
                bot.register_next_step_handler(msg, admopros)
            if call.data == 'admopros-stat':
                bot.send_message(chat_id=chat_id, text='📊 Статистика опроса.\n')
                from pycel import ExcelCompiler
                excel = ExcelCompiler(filename=localfolder + 'data/uchastok/all/opros/opros.xlsx')
                tema = f"{excel.evaluate('opros!L2')}"
                data = f"{excel.evaluate('opros!B1')}"
                area = f"{excel.evaluate('opros!B2')}"
                vsego = int(f"{excel.evaluate('opros!M7')}")
                answ0 = f"{excel.evaluate('opros!L3')}"
                answ1 = f"{excel.evaluate('opros!L4')}"
                answ2 = f"{excel.evaluate('opros!L5')}"
                result0 = int(f"{excel.evaluate('opros!M3')}")
                result1 = int(f"{excel.evaluate('opros!M4')}")
                result2 = int(f"{excel.evaluate('opros!M5')}")
                resultpercent0 = round(100 / vsego * result0, 1)
                resultpercent1 = round(100 / vsego * result1, 1)
                resultpercent2 = round(100 / vsego * result2, 1)
                vsegoround = int(round(vsego, 0))
                bot.send_message(chat_id=chat_id, text='\nДата проведения: ' + str(data) + '\nУчасток: ' + str(
                    area) + '\n\nНазвание опроса: ' + str(tema) + '\n\nРезультаты:\n\n' + str(answ0) + ' - ' + str(
                    result0) + ' | ' + str(resultpercent0) + '%' +
                                                       '\n' + str(answ1) + ' - ' + str(result1) + ' | ' + str(
                    resultpercent1) + '%' + '\n' + str(answ2) + ' - ' + str(result2) + ' | ' + str(
                    resultpercent2) + '%' + "\n\nВсего голосов: " + str(vsegoround))
                log.writelog(user_id=call.message.chat.id, username=call.from_user.first_name,
                             text=" /adm>>Опрос>>Статистика опроса ") 
            if call.data == 'admopros-clean':
                wb = openpyxl.load_workbook(localfolder + 'data/uchastok/all/opros/opros.xlsx')
                ws = wb['opros']
                ws.delete_rows(8, 500)
                ws['B2'] = "-"
                ws['M7'] = 1
                ws['M3'] = 0
                ws['M4'] = 0
                ws['M5'] = 0
                ws['L2'] = 'Опрос не начат.'
                ws['L3'] = 'Вариант №1.'
                ws['L4'] = 'Вариант №2.'
                ws['L5'] = 'Вариант №3.'
                ws['B1'] = '-'
                wb.save(localfolder + 'data/uchastok/all/opros/opros.xlsx')
                bot.send_message(chat_id=chat_id, text='📊 Результаты очищены.')
                log.writelog(user_id=call.message.chat.id, username=call.from_user.first_name,
                             text=" /adm>>Опрос>>Результаты очищены ")  
            if call.data == 'admopros-load':
                bot.send_message(chat_id=chat_id, text='🐙 Скачиваем...')
                file = open(localfolder + 'data/uchastok/all/opros/opros.xlsx', 'rb')
                bot.send_document(call.message.chat.id, file)
            if call.data == 'areammopros':
                
                bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                      text='📊 Меню работы с опросами.', parse_mode='Markdown',
                                      reply_markup=kball.areaopros)
            if call.data == 'areasend':
                
                bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id, text='Работа с рассылкой.💫',
                                      parse_mode="Markdown", reply_markup=kball.areasend)
            if call.data == 'areaopros':
                msg = bot.send_message(chat_id=chat_id,
                                       text='🐙 Показать настройки опроса?\n\n❌ стоп - для отмены\n✅ далее - для продолжения',
                                       reply_markup=kball.stopandnext)
                bot.register_next_step_handler(msg, areaopros)
            if call.data == 'areaopros-stat':
                bot.send_message(chat_id=chat_id, text='📊 Статистика опроса.\n')
                from pycel import ExcelCompiler
                excel = ExcelCompiler(filename=localfolder + 'data/uchastok/' + valid_area + '/opros/opros.xlsx')
                tema = f"{excel.evaluate('opros!L2')}"
                data = f"{excel.evaluate('opros!B1')}"
                area = f"{excel.evaluate('opros!B2')}"
                vsego = int(f"{excel.evaluate('opros!M7')}")
                answ0 = f"{excel.evaluate('opros!L3')}"
                answ1 = f"{excel.evaluate('opros!L4')}"
                answ2 = f"{excel.evaluate('opros!L5')}"
                result0 = int(f"{excel.evaluate('opros!M3')}")
                result1 = int(f"{excel.evaluate('opros!M4')}")
                result2 = int(f"{excel.evaluate('opros!M5')}")
                resultpercent0 = round(100 / vsego * result0, 1)
                resultpercent1 = round(100 / vsego * result1, 1)
                resultpercent2 = round(100 / vsego * result2, 1)
                vsegoround = int(round(vsego, 0))
                bot.send_message(chat_id=chat_id, text='\nДата проведения: ' + str(data) + '\nУчасток: ' + str(
                    area) + '\n\nНазвание опроса: ' + str(tema) + '\n\nРезультаты:\n\n' + str(answ0) + ' - ' + str(
                    result0) + ' | ' + str(resultpercent0) + '%' + '\n' + str(answ1) + ' - ' + str(
                    result1) + ' | ' + str(
                    resultpercent1) + '%' + '\n' + str(answ2) + ' - ' + str(result2) + ' | ' + str(
                    resultpercent2) + '%' + "\n\nВсего голосов: " + str(vsegoround))
                log.writelog(user_id=call.message.chat.id, username=call.from_user.first_name,
                             text=" /send>Опрос> Статистика опроса ") 
            if call.data == 'areaopros-clean':
                mycursor = mydb.cursor() 
                mydb.ping(reconnect=True)  
                mycursor.execute("SELECT area FROM users WHERE chat_id=%s",
                                 (chat_id)) 
                data = "error"  
                for i in mycursor:
                    data = i  
                if data == "error":  
                    log.writelog(user_id=call.message.chat.id, username=call.from_user.first_name,
                                 text="/send>opros>очистка БЛОКИРОВКА ПОЛЬЗОВАТЕЛЯ НЕТ В БАЗЕ >> ")
                else:
                    mycursor = mydb.cursor()
                    mycursor.execute("SELECT * FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:
                        valid_area = row['area']
                        valid_comment = row['comment']
                    mydb.commit()
                    mydb.close()
                    if valid_area == "del":
                        log.writelog_start(valid_comment=valid_comment, user_id=call.message.chat.id,
                                           username=call.from_user.first_name,
                                           text="/send>opros>очистка БЛОКИРОВКА. ПОЛЬЗОВАТЕЛЬ БЫЛ УДАЛЁН >>")
                    elif valid_area in ["mvu", "ru", "du", "tu", "svu", "sku"]:
                        wb = openpyxl.load_workbook(localfolder + 'data/uchastok/' + valid_area + '/opros/opros.xlsx')
                        ws = wb['opros']
                        ws.delete_rows(8, 500)
                        ws['B2'] = "-"
                        ws['M7'] = 1
                        ws['M3'] = 0
                        ws['M4'] = 0
                        ws['M5'] = 0
                        ws['L2'] = 'Опрос не начат.'
                        ws['L3'] = 'Вариант №1.'
                        ws['L4'] = 'Вариант №2.'
                        ws['L5'] = 'Вариант №3.'
                        ws['B1'] = '-'
                        wb.save(localfolder + 'data/uchastok/' + valid_area + '/opros/opros.xlsx')
                        bot.send_message(chat_id=chat_id, text='📊 Результаты очищены.')
                        log.writelog(user_id=call.message.chat.id, username=call.from_user.first_name,
                                     text=" /send>Опрос> Результаты очищены ")
            if call.data == 'areaopros-load':
                bot.send_message(chat_id=chat_id, text='🐙 Скачиваем...')
                file = open(localfolder + 'data/uchastok/' + valid_area + '/opros/opros.xlsx', 'rb')
                bot.send_document(call.message.chat.id, file)
                log.writelog(user_id=call.message.chat.id, username=call.from_user.first_name,
                             text="/send>opros> Скачивание опроса ") 
            if call.data == 'areastatistics':
                firstname = call.from_user.first_name  
                username = call.from_user.username  
                if username is None:
                    username = "null"
                if firstname is None:
                    firstname = "null"
                mycursor = mydb.cursor()  
                mydb.ping(reconnect=True)  
                mycursor.execute("SELECT area FROM users WHERE chat_id=%s", (chat_id))  
                data = "error"  
                for i in mycursor:
                    data = i  
                if data == "error": 
                    log.writelog(user_id=message.chat.id, username=message.from_user.first_name,
                                 text="/send>Статистика БЛОКИРОВКА ПОЛЬЗОВАТЕЛЯ НЕТ В БАЗЕ >> ") 
                else:
                    mydb.ping(
                        reconnect=True)  
                    mycursor = mydb.cursor()  
                    mycursor.execute("SELECT * FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:
                        valid_area = row['area']
                        valid_comment = row['comment']
                        valid_maillist = row['maillist']
                    mydb.commit()
                    if valid_area == "del":
                        log.writelog_start(valid_comment=valid_comment, user_id=message.chat.id,
                                           username=message.from_user.first_name,
                                           text="/send>Статистика БЛОКИРОВКА. ПОЛЬЗОВАТЕЛЬ БЫЛ УДАЛЁН >>") 
                        mydb.close()
                    elif valid_area in ["mvu", "ru", "du", "tu", "svu", "sku"]:
                        number_of_rows_all = mycursor.execute("SELECT * FROM users WHERE area=%s",
                                                              (valid_area)) 
                        number_of_rows_maillist = mycursor.execute("SELECT * FROM users WHERE area=%s and maillist=1",
                                                                   (valid_area))
                        bot.edit_message_text(chat_id=chat_id, message_id=message_id,
                                              text='Всего в участке: ' + str(
                                                  number_of_rows_all) + '\nУчаствующих в рассылке: ' + str(
                                                  number_of_rows_maillist), reply_markup=kball.areasend)
                        mydb.close()
            if call.data == 'areamessage':
                msg = bot.send_message(chat_id=chat_id,
                                       text='✏️ Введите текст для рассылки.\nДля отправки гиперссылки - [Текст кнопки](Ссылка)\n\n❌ Для отмены отправь: ❌ стоп',
                                       reply_markup=kball.stop)
                bot.register_next_step_handler(msg, areamessage1)
            if call.data == 'admusr_statistics':
                mydb.ping(
                    reconnect=True) 
                mycursor = mydb.cursor() 
                number_of_rows_all = mycursor.execute("SELECT * FROM users")
                maillist = '1'
                number_of_rows_maillist = mycursor.execute("SELECT chat_id FROM users WHERE maillist=%s", (maillist))
                bot.edit_message_text(chat_id=chat_id, message_id=message_id, text='Всего пользователей в боте: ' + str(
                    number_of_rows_all) + '\nУчаствующих в рассылке: ' + str(number_of_rows_maillist),
                                      reply_markup=kball.admusr)
                mydb.close()
                log.writelog(user_id=call.message.chat.id, username=call.from_user.first_name,
                             text=" /send Запрос статистики пользователей >>") 
            if call.data == 'admusr_message':
                msg = bot.send_message(chat_id=chat_id,
                                       text='✏️ Введите текст для рассылки.\n❌ Для отмены отправь: ❌ стоп',
                                       reply_markup=kball.stop)
                bot.register_next_step_handler(msg, admmessage1)
            if call.data == "usr_find_back":
                bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id, text='🐙 Выбери: ',
                                      reply_markup=kball.mmusr)
                
            if call.data == "usr_add":
                msg = bot.send_message(chat_id=chat_id,
                                       text='🐙 Введи ID пользователя для добавления в бота.\n❌ Для отмены отправь: стоп \n', reply_markup=kball.stop)
                bot.register_next_step_handler(msg, usr_add)
            if call.data == "usr_del":
                msg = bot.send_message(chat_id=chat_id,
                                       text='🐙 Введи ID пользователя для удаления из бота.\n❌ Для отмены отправь: стоп \n', reply_markup=kball.stop)
                bot.register_next_step_handler(msg, usr_del)
            if call.data == "usr_find":
                keyboard = types.InlineKeyboardMarkup(row_width=2)
                button1 = types.InlineKeyboardButton(text="💬 Искать по комментарию", callback_data="usr_find_comment")
                button2 = types.InlineKeyboardButton(text="🆔 Искать по ID", callback_data="usr_find_id")
                button3 = types.InlineKeyboardButton(text="📋 Список пользователей", callback_data="usr_find_all")
                button4 = types.InlineKeyboardButton(text="Назад", callback_data="usr_find_back")
                keyboard.add(button1, button2, button3, button4)
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id, text='🐙 Выбери: ',
                                                reply_markup=keyboard)
            if call.data == "usr_find_comment":
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text='🐙 Введите комментарий: ')
                bot.register_next_step_handler(in_text, usr_find_comment)
            if call.data == "usr_find_id":
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text='🐙 Введите ID: ')
                bot.register_next_step_handler(in_text, usr_find_id)
            if call.data == "usr_find_all":
                msg = bot.send_message(chat_id=chat_id, text='🐙 Список всех пользователей\n')
                chat_id = call.message.chat.id 
                mydb.ping(reconnect=True)
                mycursor = mydb.cursor()  
                mycursor.execute("SELECT area FROM users WHERE chat_id=%s",
                                 (chat_id))  
                data = "error"  
                for i in mycursor:
                    data = i  
                if data == "error":  
                    print("Нельзя")
                    mydb.close()
                else:
                    mycursor = mydb.cursor() 
                    mycursor.execute("SELECT * FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:
                        valid_area = row['area']
                    if valid_area in ["mvu", "ru", "du", "tu", "svu", "sku"]:
                        mycursor.execute("SELECT * FROM users WHERE area=%s",
                                         (valid_area))  
                        result = mycursor.fetchall()
                        for row in result:  
                            find_chat_id = row["chat_id"]
                            find_nick = row["nick"]
                            find_comment = row["comment"]
                            msg = bot.send_message(chat_id=chat_id, text="🆔: " + str(find_chat_id) + "\nНик: @" + str(
                                find_nick) + "\nИмя: " + str(find_comment))
                            time.sleep(0.05) 
                    msg = bot.send_message(chat_id=chat_id, text='---------------', reply_markup=kball.mmusr)
                    mydb.close()  
            if call.data == "adm_main":
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text="🐙 Меню административное.💫", reply_markup=kball.mmadm)
                
            if call.data == "adm_main_ut":
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text="🐙 Выбери станцию: ", reply_markup=kball.adm_ut)
                
            if call.data == "adm_info":
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text="🐙 Выбери раздел: ", reply_markup=kball.adm_info)
            if call.data == "adm_ut":
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text="🐙 Выбери станцию: ", reply_markup=kball.adm_ut)
            if call.data == "adm_user":
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text="🐙 Управление пользователями: ", reply_markup=kball.admusr)
            if call.data == "mmadmutmv":
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text="🐙\n⚠️Пожалуйста, будь аккуратен.\nПерезагрузка реле не спрашивает подтверждения.",
                                                reply_markup=kball.mmadmutmv)
            if call.data == "mmadmut3km":
              
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text="🐙\n⚠️Пожалуйста, будь аккуратен.\nПерезагрузка реле не спрашивает подтверждения.",
                                                reply_markup=kball.mmadmut3km)
            if call.data == "mmadmutinoz":
                
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text="🐙\n⚠️Пожалуйста, будь аккуратен.\nПерезагрузка реле не спрашивает подтверждения.",
                                                reply_markup=kball.mmadmutinoz)
            if call.data == "mmadmutlermont":
               
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text="🐙\n⚠️Пожалуйста, будь аккуратен.\nПерезагрузка реле не спрашивает подтверждения.",
                                                reply_markup=kball.mmadmutlermont)
            if call.data == "mmadmutminutka":
                
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text="🐙\n⚠️Пожалуйста, будь аккуратен.\nПерезагрузка реле не спрашивает подтверждения.",
                                                reply_markup=kball.mmadmutminutka)
            if call.data == "adm_logs":
                
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text="🐙 Выбери объём логов: ", reply_markup=kball.adm_logs)
            if call.data == "logs_send_20":
                from subprocess import check_output
                bot.send_message(call.message.chat.id,
                                 check_output('tail -n 20 ' + localfolder + 'log/bot.log', shell=True),
                                 reply_markup=kball.adm_logs)
                log.writelog(user_id=call.from_user.id, username=call.from_user.username,
                             text="Запрос логов 20")  
            if call.data == "logs_send_50":
                from subprocess import check_output
                bot.send_message(call.message.chat.id,
                                 check_output('tail -n 50 ' + localfolder + 'log/bot.log', shell=True),
                                 reply_markup=kball.adm_logs)
                log.writelog(user_id=call.from_user.id, username=call.from_user.username,
                             text="Запрос логов 50")  
            if call.data == "adm_ofd":
                
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text="🐙 Меню работы с ОФД: ", reply_markup=kball.adm_ofd)
            if call.data == "ofd_report_fd":
                mycursordate = mydb.cursor() 
                mydb.ping(reconnect=True)
                mycursordate.execute("SELECT datereq FROM ofd")
                resultdate = mycursordate.fetchone()
                mycursor = mydb.cursor()  
                mydb.ping(reconnect=True)
                mycursor.execute("SELECT * FROM ofd WHERE fd > 200000")
                result = mycursor.fetchall()
                mydb.close()
                for row in result:
                    kkt = row['kkt']
                    fn = row['fn']
                    fd = row['fd']
                    fnend = row['fnend']
                    place = row['place']
                    depart = row['depart']
                    check_day_all = row['check_day_all']
                    possibledayfdend = row['possibledayfdend']
                    bot.send_message(chat_id, '🧾ФД: *' + str(fd) + '*\nЧеков за день: ' + str(check_day_all) + '\nХватит на ~' + str(possibledayfdend) + ' дней\nККТ: ' + str(kkt) + "\nФН: " + str(fn) + "\nСрок действия ФН: " + str(fnend) + "\nТорговая точка: *" + str(place) + "*\nПодразделение: " + str(depart), parse_mode="Markdown")
                bot.send_message(chat_id, '🐙 Отчет сформирован за ' + str(
                    resultdate['datereq']) + '\nСписок ФН с количеством ФД выше 200к.\n', parse_mode="Markdown", reply_markup=kball.adm_ofd)
                log.writelog(user_id=call.from_user.id, username=call.from_user.username,
                             text="ОФД. Отчёт по ФД") 
            if call.data == "ofd_report_fn":
                from datetime import date
                from datetime import timedelta
                date_request30 = str(date.today() + timedelta(days=+30))  
                date_request = str(date.today() + timedelta(days=-1))  
                today = date.today()
                mycursordate = mydb.cursor()  
                mydb.ping(reconnect=True)
                mycursordate.execute("SELECT datereq FROM ofd")
                resultdate = mycursordate.fetchone()
                mycursor.execute("SELECT * FROM ofd WHERE fnend BETWEEN '" + str(today) + "' AND '" + str(
                    date_request30) + "' ORDER BY fnend DESC")
                mydb.ping(reconnect=True)
                result = mycursor.fetchall()
                mydb.close()
                for row in result:
                    kkt = row['kkt']
                    fn = row['fn']
                    fd = row['fd']
                    fnend = row['fnend']
                    place = row['place']
                    depart = row['depart']
                    bot.send_message(chat_id, '🕑ККТ: *' + str(kkt) + '*\nФН: ' + str(
                        fn) + "\nФД: " + str(fd) + "\nСрок действия ФН: *" + str(fnend) + "*\nТорговая точка: *" + str(
                        place) + "*\nПодразделение: " + str(depart), parse_mode="Markdown")
                bot.send_message(chat_id, '🐙 Отчет сформирован за ' + str(
                    resultdate['datereq']) + '\nСписок ФН которые заканчиваются в ближайшие 30 дней.\n', parse_mode="Markdown",
                                 reply_markup=kball.adm_ofd)




                log.writelog(user_id=call.from_user.id, username=call.from_user.username,
                             text="ОФД. Отчёт по ФН")  
            if call.data == "ofd_report_sub":
                chat_id = call.message.chat.id
                firstname = call.from_user.first_name
                username = call.from_user.username
                mydb.ping(
                    reconnect=True) 
                mycursor = mydb.cursor()  
                mycursor.execute("SELECT * FROM users WHERE chat_id=%s", (chat_id))
                result = mycursor.fetchall()
                for row in result:

                    find_area = row['area']
                    find_comment = row['comment']
                    find_maillist_ofd = row['maillist_ofd']
                    if find_area == "mvu":
                        find_area = "МВУ"
                    elif find_area == "ru":
                        find_area = "РУ"
                    elif find_area == "du":
                        find_area = "ДУ"
                    elif find_area == "tu":
                        find_area = "ТУ"
                    elif find_area == "svu":
                        find_area = "СВУ"
                    elif find_area == "sku":
                        find_area = "СКУ"
                    elif find_area == "del":
                        find_area = "Заблокирован"
                    if find_maillist_ofd == 1:
                        find_maillist_ofd = "✅ Подписан"
                        action_maillist_ofd = "🛑 Отписаться"
                    else:
                        find_maillist_ofd = "🛑 Не подписан"
                        action_maillist_ofd = "✅ Подписаться"
                mydb.commit()
                mydb.close()
                set_sub = types.InlineKeyboardMarkup(row_width=1)
                button1 = types.InlineKeyboardButton(str(action_maillist_ofd),
                                                     callback_data='call_sub_maillist_ofd')
                button2 = types.InlineKeyboardButton("⬅ назад  ", callback_data='mainmenu_ofd')
                set_sub.add(button1, button2)
                bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                      text="🐙 Активировать ежедневное информирование о критическом состоянии ОФД (За 14 дней до истечения ФН. Переполнение ФД свыше 240к).\n*Данные обновляются в 09:00*\n\nО пользователе:\n👤ФИО: " + find_comment + "\n🚈 Участок: " + find_area + "\n📤 Рассылка ОФД: " + str(
                                          find_maillist_ofd),
                                      parse_mode="Markdown", reply_markup=set_sub)

                log.writelog(user_id=call.from_user.id, username=call.from_user.username,
                             text="ОФД. Подписка\отписка") 
            if call.data == "call_sub_maillist_ofd":
                chat_id = call.message.chat.id
                firstname = call.from_user.first_name
                username = call.from_user.username
                mydb.ping(reconnect=True) 
                mycursor = mydb.cursor() 
                mycursor.execute("SELECT * FROM users WHERE chat_id=%s", (chat_id))
                result = mycursor.fetchall()
                for row in result:  
                    find_maillist_ofd = row["maillist_ofd"]
                    if find_maillist_ofd == 1:
                        mycursor.execute("UPDATE users SET maillist_ofd = '0' WHERE chat_id=%s", (chat_id))
                        comment = "🛑 Уведомления отключены"
                        mydb.commit() 
                    else:
                        mycursor.execute("UPDATE users SET maillist_ofd = '1' WHERE chat_id=%s", (chat_id)) 
                        comment = "✅ Уведомления включены"
                        mydb.commit()  

                bot.send_message(chat_id=chat_id, text='🐙 Рассылка ОФД- ' + comment)
                mydb.close() 
            if call.data == "mainmenu_ofd":
                chat_id = call.message.chat.id
                firstname = call.from_user.first_name
                username = call.from_user.username
                bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id, text="🐙 Меню работы с ОФД:", parse_mode="Markdown", reply_markup=kball.adm_ofd)

            if call.data == "adm_redmine":
                
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text="🐙 Меню Redmine: ", reply_markup=kball.adm_redmine)
            if call.data == "adm_settings":
                mydb.ping(reconnect=True)
                mycursor = mydb.cursor()  
                mycursor.execute("SELECT redminelogindb FROM users WHERE chat_id=%s", (chat_id))
                result = mycursor.fetchall()
                for row in result: 
                    redminename = row["redminelogindb"]
                mydb.close() 
                log.writelog(user_id=call.from_user.id, username=call.from_user.username,
                             text="/adm Настройки") 
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text="🐙 Меню настроек:\nRedmine логин: " + str(redminename),
                                                reply_markup=kball.adm_settings)
            if call.data == 'redmine_auth_in':
                msg = bot.send_message(call.message.chat.id, "🐙 Введи Логин Redmine: ")
                bot.register_next_step_handler(msg, process_redminelogin)
            if call.data == 'redmine_auth_out':
                rname = 0
                rpass = 0
                bot.send_message(call.message.chat.id, "🐙 Выходим из аккаунта Redmine...")
                mydb.ping(reconnect=True)
                mycursor = mydb.cursor()  
                mycursor.execute("UPDATE users set redminelogindb = %s WHERE chat_id=%s", (rname, chat_id))
                mycursor.execute("UPDATE users set redminepassdb = %s WHERE chat_id=%s", (rpass, chat_id))
                mydb.commit()
                mydb.close()
                bot.send_message(call.message.chat.id, "🐙 Успешно!", reply_markup=kball.adm_settings)
            if call.data == "redmine_mm":
                
                bot.send_message(call.message.chat.id, "️️🐙 Назад", reply_markup=kball.adm_redmine)
            if call.data == "redmine_mk":
                
                bot.send_message(call.message.chat.id, "️️🐙 Мобильные кассы. Выбери период.",
                                 reply_markup=kball.redminemk)
            if call.data == "redmine_mk_7":
                try:
                    mydb.ping(reconnect=True)
                    mycursor = mydb.cursor()  
                    chat_id = call.message.chat.id
                    mycursor.execute("SELECT redminelogindb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:  
                        a = row["redminelogindb"]

                    mycursor.execute("SELECT redminepassdb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:  
                        b = row["redminepassdb"]
                    mydb.commit()
                    mydb.close()
                    redmine = Redmine('url_redmine', username=a, password=b)
                    last10redmine = list(redmine.issue.filter(status_id=('o'), tracker_id=('22'),
                                                              created_on='>=%s' % checktime7.strftime('%Y-%m-%d')))
                    for issue in last10redmine:
                        url = 'url_redmineissues/'
                        bot.send_message(call.message.chat.id,
                                         url + str(issue.id) + "\n" + str(issue) + "\nСоздано: " + str(
                                             issue.start_date) + '\nНазначено: ' + str(issue.assigned_to))
                    bot.send_message(call.message.chat.id,
                                     "️️🐙 Мобильные кассы. Список (открытых\в работе) заявок за 7 дней",
                                     reply_markup=kball.redminemk)
                    log.writelog(user_id=call.from_user.id, username=call.from_user.username,
                                 text="Redmine. Мобильные кассы. Список (открытых\в работе) заявок за 7 дней") 
                except Exception as e:
                    print(e)
            if call.data == "redmine_mk_14":
                try:
                    mydb.ping(reconnect=True)
                    mycursor = mydb.cursor()  
                    chat_id = call.message.chat.id
                    mycursor.execute("SELECT redminelogindb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result: 
                        a = row["redminelogindb"]

                    mycursor.execute("SELECT redminepassdb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:  
                        b = row["redminepassdb"]
                    mydb.commit()
                    mydb.close()
                    redmine = Redmine('url_redmine', username=a, password=b)
                    last10redmine = list(redmine.issue.filter(status_id=('o'), tracker_id=('22'),
                                                              created_on='>=%s' % checktime14.strftime('%Y-%m-%d')))
                    for issue in last10redmine:
                        url = 'url_redmineissues/'
                        bot.send_message(call.message.chat.id,
                                         url + str(issue.id) + "\n" + str(issue) + "\nСоздано: " + str(
                                             issue.start_date) + '\nНазначено: ' + str(issue.assigned_to))
                    bot.send_message(call.message.chat.id,
                                     "️️🐙 Мобильные кассы. Список (открытых\в работе) заявок за 14 дней",
                                     reply_markup=kball.redminemk)
                    log.writelog(user_id=call.from_user.id, username=call.from_user.username,
                                 text="Redmine. Мобильные кассы. Список (открытых\в работе) заявок за 14 дней")  
                except Exception as e:
                    bot.send_message(call.message.chat.id, '🐙 Не указана авторизация')
            if call.data == "redmine_apm":
                
                bot.send_message(call.message.chat.id, "️️🐙 Стационарные кассы. Выбери период.",
                                 reply_markup=kball.redmineapm)
            if call.data == "redmine_apm_7":
                try:
                    mydb.ping(reconnect=True)
                    mycursor = mydb.cursor()  
                    chat_id = call.message.chat.id
                    mycursor.execute("SELECT redminelogindb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:  
                        a = row["redminelogindb"]

                    mycursor.execute("SELECT redminepassdb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:  
                        b = row["redminepassdb"]
                    mydb.commit()
                    mydb.close()
                    redmine = Redmine('url_redmine', username=a, password=b)
                    last10redmine = list(redmine.issue.filter(status_id=('o'), tracker_id=('7'),
                                                              created_on='>=%s' % checktime7.strftime('%Y-%m-%d')))
                    for issue in last10redmine:
                        url = 'url_redmineissues/'
                        bot.send_message(call.message.chat.id,
                                         url + str(issue.id) + "\n" + str(issue) + "\nСоздано: " + str(
                                             issue.start_date) + '\nНазначено: ' + str(issue.assigned_to))
                    bot.send_message(call.message.chat.id,
                                     "️️🐙 Стационарные кассы. Список (открытых\в работе) заявок за 7 дней",
                                     reply_markup=kball.redmineapm)
                    log.writelog(user_id=call.from_user.id, username=call.from_user.username,
                                 text="Redmine. Стационарные кассы. Список (открытых\в работе) заявок за 7 дней") 
                except Exception as e:
                    bot.send_message(call.message.chat.id, '🐙 Не указана авторизация')
            if call.data == "redmine_apm_14":
                try:
                    mydb.ping(reconnect=True)
                    mycursor = mydb.cursor()  
                    chat_id = call.message.chat.id
                    mycursor.execute("SELECT redminelogindb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:  
                        a = row["redminelogindb"]

                    mycursor.execute("SELECT redminepassdb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:  
                        b = row["redminepassdb"]
                    mydb.commit()
                    mydb.close()
                    redmine = Redmine('url_redmine', username=a, password=b)
                    last10redmine = list(redmine.issue.filter(status_id=('o'), tracker_id=('7'),
                                                              created_on='>=%s' % checktime14.strftime('%Y-%m-%d')))
                    for issue in last10redmine:
                        url = 'url_redmineissues/'
                        bot.send_message(call.message.chat.id,
                                         url + str(issue.id) + "\n" + str(issue) + "\nСоздано: " + str(
                                             issue.start_date) + '\nНазначено: ' + str(issue.assigned_to))
                    bot.send_message(call.message.chat.id,
                                     "️️🐙 Стационарные кассы. Список (открытых\в работе) заявок за 14 дней",
                                     reply_markup=kball.redmineapm)
                    log.writelog(user_id=call.from_user.id, username=call.from_user.username,
                                 text="Redmine. Стационарные кассы. Список (открытых\в работе) заявок за 14 дней")  
                except Exception as e:
                    bot.send_message(call.message.chat.id, '🐙 Не указана авторизация')
            if call.data == "redmine_bpa":
                
                bot.send_message(call.message.chat.id, "️️🐙 Терминалы самообслуживания. Выбери период.",
                                 reply_markup=kball.redminebpa)
            if call.data == "redmine_bpa_7":
                try:
                    mydb.ping(reconnect=True)
                    mycursor = mydb.cursor()  
                    chat_id = call.message.chat.id
                    mycursor.execute("SELECT redminelogindb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:  
                        a = row["redminelogindb"]

                    mycursor.execute("SELECT redminepassdb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:  
                        b = row["redminepassdb"]
                    mydb.commit()
                    mydb.close()
                    redmine = Redmine('url_redmine', username=a, password=b)
                    last10redmine = list(redmine.issue.filter(status_id=('o'), tracker_id=('10'),
                                                              created_on='>=%s' % checktime7.strftime('%Y-%m-%d')))
                    for issue in last10redmine:
                        url = 'url_redmineissues/'
                        bot.send_message(call.message.chat.id,
                                         url + str(issue.id) + "\n" + str(issue) + "\nСоздано: " + str(
                                             issue.start_date) + '\nНазначено: ' + str(issue.assigned_to))
                    bot.send_message(call.message.chat.id,
                                     "️️🐙 Терминалы самообслуживания. Список (открытых\в работе) заявок за 7 дней",
                                     reply_markup=kball.redminebpa)
                    log.writelog(user_id=call.from_user.id, username=call.from_user.username,
                                 text="Redmine. Терминалы самообслуживания. Список (открытых\в работе) заявок за 7 дней")  
                except Exception as e:
                    bot.send_message(call.message.chat.id, '🐙 Не указана авторизация')
            if call.data == "redmine_bpa_14":
                try:
                    mydb.ping(reconnect=True)
                    mycursor = mydb.cursor()  
                    chat_id = call.message.chat.id
                    mycursor.execute("SELECT redminelogindb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:  
                        a = row["redminelogindb"]

                    mycursor.execute("SELECT redminepassdb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result: 
                        b = row["redminepassdb"]
                    mydb.commit()
                    mydb.close()
                    redmine = Redmine('url_redmine', username=a, password=b)
                    last10redmine = list(redmine.issue.filter(status_id=('o'), tracker_id=('10'),
                                                              created_on='>=%s' % checktime14.strftime('%Y-%m-%d')))
                    for issue in last10redmine:
                        url = 'url_redmineissues/'
                        bot.send_message(call.message.chat.id,
                                         url + str(issue.id) + "\n" + str(issue) + "\nСоздано: " + str(
                                             issue.start_date) + '\nНазначено: ' + str(issue.assigned_to))
                    bot.send_message(call.message.chat.id,
                                     "️️🐙 Терминалы самообслуживания. Список (открытых\в работе) заявок за 14 дней",
                                     reply_markup=kball.redminebpa)
                    log.writelog(user_id=call.from_user.id, username=call.from_user.username,
                                 text="Redmine. Терминалы самообслуживания. Список (открытых\в работе) заявок за 14 дней") 
                except Exception as e:
                    bot.send_message(call.message.chat.id, '🐙 Не указана авторизация')
            if call.data == "redmine_ut":
                bot.send_message(call.message.chat.id, "️️🐙 Список открытых заявок. Выбери период.",
                                 reply_markup=kball.redmineutopen)
                bot.send_message(call.message.chat.id, "️️🐙 Список закрытых заявок. Выбери период.",
                                 reply_markup=kball.redmineutclose)
            if call.data == "redmine_utop_7":
                try:
                    mydb.ping(reconnect=True)
                    mycursor = mydb.cursor()  
                    chat_id = call.message.chat.id
                    mycursor.execute("SELECT redminelogindb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result: 
                        a = row["redminelogindb"]

                    mycursor.execute("SELECT redminepassdb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result: 
                        b = row["redminepassdb"]
                    mydb.commit()
                    mydb.close()
                    redmine = Redmine('url_redmine', username=a, password=b)
                    issues = list(redmine.issue.filter(status_id=('o'), tracker_id=('21'),
                                                       created_on='>=%s' % checktime7.strftime('%Y-%m-%d')))
                    for issue in issues:
                        url = 'url_redmineissues/'
                        bot.send_message(call.message.chat.id,
                                         url + str(issue.id) + "\n" + str(issue) + "\nСоздано: " + str(
                                             issue.start_date) + '\nНазначено: ' + str(issue.assigned_to))
                    bot.send_message(call.message.chat.id, "️️🐙 Список (открытых\в работе) заявок за 7 дней",
                                     reply_markup=kball.adm_redmine)
                    log.writelog(user_id=call.from_user.id, username=call.from_user.username,
                                 text="Redmine. Список (открытых\в работе) заявок за 7 дней")  
                except Exception as e:
                    bot.send_message(call.message.chat.id, '🐙 Не указана авторизация')
            if call.data == "redmine_utcl_30":
                try:
                    mydb.ping(reconnect=True)
                    mycursor = mydb.cursor() 
                    chat_id = call.message.chat.id
                    mycursor.execute("SELECT redminelogindb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result: 
                        a = row["redminelogindb"]

                    mycursor.execute("SELECT redminepassdb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:  
                        b = row["redminepassdb"]
                    mydb.commit()
                    mydb.close()
                    redmine = Redmine('url_redmine', username=a, password=b)
                    issues = list(redmine.issue.filter(status_id=('c'), tracker_id=('21'),
                                                       created_on='>=%s' % checktime30.strftime('%Y-%m-%d')))
                    for issue in issues:
                        url = 'url_redmineissues/'
                        bot.send_message(call.message.chat.id,
                                         url + str(issue.id) + "\n" + str(issue) + "\nСоздано: " + str(
                                             issue.start_date) + '\nНазначено: ' + str(issue.assigned_to))
                    bot.send_message(call.message.chat.id, "️️🐙 Список закрытых заявок за 30 дней",
                                     reply_markup=kball.adm_redmine)
                    log.writelog(user_id=call.from_user.id, username=call.from_user.username,
                                 text="Redmine. Список закрытых заявок за 30 дней") 
                except Exception as e:
                    bot.send_message(call.message.chat.id, '🐙 Не указана авторизация')
            if call.data == "redmine_all":
                bot.send_message(call.message.chat.id, "️️🐙 Список открытых заявок. Выбери период.",
                                 reply_markup=kball.redmineall)
                bot.send_message(call.message.chat.id, "️️🐙 Список всех заявок. Выбери период.",
                                 reply_markup=kball.redmineallfull)
            if call.data == "redmine_all_7":
                try:
                    mydb.ping(reconnect=True)
                    mycursor = mydb.cursor() 
                    chat_id = call.message.chat.id
                    mycursor.execute("SELECT redminelogindb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:  
                        a = row["redminelogindb"]

                    mycursor.execute("SELECT redminepassdb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:  
                        b = row["redminepassdb"]
                    mydb.commit()
                    mydb.close()
                    redmine = Redmine('url_redmine', username=a, password=b)
                    issues = list(redmine.issue.filter(project_id=('skppk'), subproject_id=('!*'),
                                                       created_on='>=%s' % checktime7.strftime('%Y-%m-%d')))
                    for issue in issues:
                        url = 'url_redmineissues/'
                        bot.send_message(call.message.chat.id,
                                         url + str(issue.id) + "\n" + str(issue) + "\nСоздано: " + str(
                                             issue.start_date) + '\nНазначено: ' + str(issue.assigned_to))
                    bot.send_message(call.message.chat.id, "️️🐙 Список (открытых\в работе) заявок за 7 дней",
                                     reply_markup=kball.adm_redmine)
                    log.writelog(user_id=call.from_user.id, username=call.from_user.username,
                                 text="Redmine. Список (открытых\в работе) заявок за 7 дней") 
                except Exception as e:
                    bot.send_message(call.message.chat.id, '🐙 Не указана авторизация')
           
            if call.data == "redmine_my":
                try:
                    mydb.ping(reconnect=True)
                    mycursor = mydb.cursor() 
                    chat_id = call.message.chat.id
                    mycursor.execute("SELECT redminelogindb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:  
                        a = row["redminelogindb"]

                    mycursor.execute("SELECT redminepassdb FROM users WHERE chat_id=%s", (chat_id))
                    result = mycursor.fetchall()
                    for row in result:  
                        b = row["redminepassdb"]
                    mydb.commit()
                    mydb.close()
                    redmine = Redmine('url_redmine', username=a, password=b)
                    issues = list(redmine.issue.filter(status_id=('o'), assigned_to_id=('me')))
                    for issue in issues:
                        url = 'url_redmineissues/'
                        bot.send_message(call.message.chat.id,
                                         url + str(issue.id) + "\n" + str(issue) + "\nСоздано: " + str(
                                             issue.start_date) + '\nНазначено: ' + str(issue.assigned_to))
                    bot.send_message(call.message.chat.id, "️️🐙 Список назначенных заявок: ",
                                     reply_markup=kball.adm_redmine)
                    log.writelog(user_id=call.from_user.id, username=call.from_user.username,
                                 text="Redmine. Список назначенных заявок мне")
                except Exception as e:
                    bot.send_message(call.message.chat.id, '🐙 Не указана авторизация')
            if call.data == "adm_user":
                
                bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                      text="Управление пользователями", reply_markup=kball.admusr)
            if call.data == "admusr":
                
                bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                      text="Управление пользователями", reply_markup=kball.admusr)
            if call.data == "admusr_add":
                msg = bot.send_message(chat_id=chat_id,
                                       text='🐙 Введи ID пользователя для добавления в бота.\n❌ Для отмены отправь: стоп \n')
                bot.register_next_step_handler(msg, admusr_add)
            if call.data == "admusr_del":
                msg = bot.send_message(chat_id=chat_id,
                                       text='🐙 Введи ID пользователя для удаления из бота.\n❌ Для отмены отправь: стоп \n')
                bot.register_next_step_handler(msg, admusr_del)
            if call.data == "admusr_charea":
                msg = bot.send_message(chat_id=chat_id,
                                       text='🐙 Введи ID пользователя для изменения участка.\n❌ Для отмены отправь: ❌ стоп',
                                       reply_markup=kball.stop)
                bot.register_next_step_handler(msg, admusr_charea1)
            if call.data == "admall_cam":
                
                bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id, text="Управление 📸 cam",
                                      reply_markup=kball.admusrcam)
            if call.data == "admall_adm":
                
                bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id, text="Управление 🐙 adm",
                                      reply_markup=kball.admusradm)
            if call.data == "admall_usr":
                
                bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id, text="Управление 👤 usr",
                                      reply_markup=kball.admusrusr)
            if call.data == "admall_send":
                
                bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id, text="Управление ✉ send",
                                      reply_markup=kball.admusrsend)
            if call.data == "admusr_list_cam":
                chat_id = call.message.chat.id  
                firstname = call.from_user.first_name  
                msg = bot.send_message(chat_id=chat_id, text='🐙 Пользователи с доступом в 📸 cam \n')
                mycursor = mydb.cursor()  
                mydb.ping(
                    reconnect=True) 
                cam = "1"
                mycursor.execute("SELECT * FROM users WHERE cam=%s", (cam))  
                result = mycursor.fetchall()
                for row in result: 
                    find_chat_id = row["chat_id"]
                    find_nick = row["nick"]
                    find_comment = row["comment"]
                    msg = bot.send_message(chat_id=chat_id, text="🆔: " + str(find_chat_id) + "\nНик: @" + str(
                        find_nick) + "\nИмя: " + str(find_comment))
                mydb.close()  
                msg = bot.send_message(chat_id=chat_id, text='-----', reply_markup=kball.admusrcam)
                msg = bot.send_message(admin, text='🔎 Выполнен поиск всех пользователей в разделе /cam\nID:  ' + str(chat_id) + '\nПользователь: ' + firstname)
                log.writelog(user_id=call.message.chat.id, username=call.from_user.first_name,
                             text="Выполнен поиск всех пользователей /cam >> ") 
            if call.data == "admusr_add_cam":
                msg = bot.send_message(chat_id=chat_id,
                                       text='🐙 Введи ID пользователя для добавления в 📸 cam.\n❌ Для отмены отправь: стоп \n')
                bot.register_next_step_handler(msg, admusr_add_cam)
            if call.data == "admusr_del_cam":
                msg = bot.send_message(chat_id=chat_id,
                                       text='🐙 Введи ID пользователя для удаления из 📸 cam.\n❌ Для отмены отправь: стоп \n')
                bot.register_next_step_handler(msg, admusr_del_cam)
            if call.data == "admusr_list_adm":
                chat_id = call.message.chat.id 
                firstname = call.from_user.first_name 
                msg = bot.send_message(chat_id=chat_id, text='🐙 Пользователи с доступом в 🐙 adm')
                mycursor = mydb.cursor()  
                mydb.ping(
                    reconnect=True)  
                adm = "1"
                mycursor.execute("SELECT * FROM users WHERE adm=%s", (adm))  
                result = mycursor.fetchall()
                for row in result:  
                    find_chat_id = row["chat_id"]
                    find_nick = row["nick"]
                    find_comment = row["comment"]
                    msg = bot.send_message(chat_id=chat_id, text="🆔: " + str(find_chat_id) + "\nНик: @" + str(
                        find_nick) + "\nИмя: " + str(find_comment))
                mydb.close()  
                msg = bot.send_message(chat_id=chat_id, text='-----', reply_markup=kball.admusradm)
                msg = bot.send_message(admin, text='🔎 Выполнен поиск всех пользователей в разделе /adm\nID:  ' + str(
                    chat_id) + '\nПользователь: ' + firstname)
                log.writelog(user_id=call.message.chat.id, username=call.from_user.first_name,
                             text="Выполнен поиск всех пользователей /adm >> ") 
            if call.data == "admusr_add_adm":
                
                msg = bot.send_message(chat_id=chat_id, text='🚫Упс... Нет такой возможности.')
            if call.data == "admusr_del_adm":
                
                msg = bot.send_message(chat_id=chat_id, text='🚫Упс... Нет такой возможности.')
            if call.data == "admusr_list_usr":
                chat_id = call.message.chat.id 
                firstname = call.from_user.first_name 
                msg = bot.send_message(chat_id=chat_id, text='🐙 Пользователи с доступом в 👤 usr')
                mycursor = mydb.cursor()  
                mydb.ping(
                    reconnect=True)  
                usr = "1"
                mycursor.execute("SELECT * FROM users WHERE usr=%s", (usr))  
                result = mycursor.fetchall()
                for row in result:  
                    find_chat_id = row["chat_id"]
                    find_nick = row["nick"]
                    find_comment = row["comment"]
                    msg = bot.send_message(chat_id=chat_id, text="🆔: " + str(find_chat_id) + "\nНик: @" + str(
                        find_nick) + "\nИмя: " + str(find_comment))
                mydb.close()
                msg = bot.send_message(chat_id=chat_id, text='-----', reply_markup=kball.admusrusr)
                msg = bot.send_message(admin, text='🔎 Выполнен поиск всех пользователей в разделе /usr\nID:  ' + str(
                    chat_id) + '\nПользователь: ' + firstname)
                log.writelog(user_id=call.message.chat.id, username=call.from_user.first_name,
                             text="Выполнен поиск всех пользователей /usr >> ")  
            if call.data == "admusr_add_usr":
                msg = bot.send_message(chat_id=chat_id,
                                       text='🐙 Введи ID пользователя для добавления в 👤 usr.\n❌ Для отмены отправь: стоп \n')
                bot.register_next_step_handler(msg, admusr_add_usr)
            if call.data == "admusr_del_usr":
                msg = bot.send_message(chat_id=chat_id,
                                       text='🐙 Введи ID пользователя для удаления из 👤 usr.\n❌ Для отмены отправь: стоп \n')
                bot.register_next_step_handler(msg, admusr_del_usr)
            if call.data == "admusr_list_send":
                chat_id = call.message.chat.id 
                firstname = call.from_user.first_name  
                msg = bot.send_message(chat_id=chat_id, text='🐙 Пользователи с доступом в ✉ send')
                mycursor = mydb.cursor()  
                mydb.ping(
                    reconnect=True)  
                send = "1"
                mycursor.execute("SELECT * FROM users WHERE send=%s", (send)) 
                result = mycursor.fetchall()
                for row in result:  
                    find_chat_id = row["chat_id"]
                    find_nick = row["nick"]
                    find_comment = row["comment"]
                    msg = bot.send_message(chat_id=chat_id, text="🆔: " + str(find_chat_id) + "\nНик: @" + str(
                        find_nick) + "\nИмя: " + str(find_comment))
                mydb.close()  
                msg = bot.send_message(chat_id=chat_id, text='-----', reply_markup=kball.admusrsend)
                msg = bot.send_message(admin, text='🔎 Выполнен поиск всех пользователей в разделе /send\nID:  ' + str(
                    chat_id) + '\nПользователь: ' + firstname)
                log.writelog(user_id=call.message.chat.id, username=call.from_user.first_name,
                             text="Выполнен поиск всех пользователей /send >> ")  
            if call.data == "admusr_add_send":
                msg = bot.send_message(chat_id=chat_id,
                                       text='🐙 Введи ID пользователя для добавления в ✉ send.\n❌ Для отмены отправь: стоп \n')
                bot.register_next_step_handler(msg, admusr_add_send)
            if call.data == "admusr_del_send":
                msg = bot.send_message(chat_id=chat_id,
                                       text='🐙 Введи ID пользователя для добавления в ✉ send.\n❌ Для отмены отправь: стоп \n')
                bot.register_next_step_handler(msg, admusr_del_send)
            if call.data == "admusr_find":
                keyboard = types.InlineKeyboardMarkup(row_width=2)
                button1 = types.InlineKeyboardButton(text="💬 Искать по комментарию",
                                                     callback_data="admusr_find_comment")
                button2 = types.InlineKeyboardButton(text="🆔 Искать по ID", callback_data="admusr_find_id")
                button3 = types.InlineKeyboardButton(text="Назад", callback_data="admusr_find_back")
                keyboard.add(button1, button2, button3)
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id, text='🐙 Выбери: ',
                                                reply_markup=keyboard)
            if call.data == "admusr_find_back":
                
                bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id, text='🐙 Выбери: ',
                                      reply_markup=kball.admusr)
            if call.data == "admusr_find_comment":
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text='🐙 Введите комментарий: ')
                bot.register_next_step_handler(in_text, admusr_find_comment)
            if call.data == "admusr_find_id":
                in_text = bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id,
                                                text='🐙 Введите ID: ')
                bot.register_next_step_handler(in_text, admusr_find_id)
            if call.data == "howto":
                file = open(localfolder + 'data/uchastok/all/howto.gif', 'rb')
                bot.send_document(call.message.chat.id, file)
            if call.data == "mm_zayavka":
                
                bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id, text="🐙 Воспользуйся кнопкой 🚨Помощь,\n там точно  есть решение.\n Если не нашёл ответ или он не помог - воспользуйся формой: 📨 Заполнить заявку!", reply_markup=kball.mmzayav)
            if call.data == "tel_req":
                chat_id = call.message.chat.id
                firstname = call.from_user.first_name
                username = call.from_user.username
                mydb.ping(
                    reconnect=True)  
                mycursor = mydb.cursor() 
                mycursor.execute("SELECT * FROM users WHERE chat_id=%s", (chat_id))
                result = mycursor.fetchall()
                for row in result:
                    valid_area = row['area']
                    valid_comment = row['comment']
                mydb.commit()
                mydb.close()
                if valid_area in ["mvu", "ru", "du", "tu", "svu", "sku"]:
                    bot.send_message(call.message.chat.id, "🐙 Номера телефонов и график работы:\n")
                    file = open(localfolder + 'data/uchastok/' + valid_area + '/documents/grafiki/' + valid_area  + '_sbk.jpg', 'rb')
                    bot.send_photo(call.message.chat.id, file)
            if call.data == "zapolnit_zayavka":
                msg = bot.send_message(call.message.chat.id,
                                       "🐙 Я помогу тебе подготовить заявку.\n Ответь на несколько вопросов:\n\n1️⃣ Отправь мне ФИО")
                bot.register_next_step_handler(msg, process_name_step)
            if call.data == "mm_all_bank":
                bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id, text="🐙 Есть идея или предложения? Отлично!\n Можешь внести свой вклад в наше развитие.\n Воспользуйся формой ниже: ", reply_markup=kball.mmbankin)
            if call.data == "bank_suggest":
                msg = bot.send_message(call.message.chat.id, "🐙 Я помогу тебе подготовить предложение.\n Выбери область применения:", reply_markup=kball.bankquestion)
                bot.register_next_step_handler(msg, process_bank1_step)
            if call.data == "usr_settings":
                chat_id = call.message.chat.id
                firstname = call.from_user.first_name
                username = call.from_user.username
                mydb.ping(
                    reconnect=True) 
                mycursor = mydb.cursor() 
                mycursor.execute("SELECT * FROM users WHERE chat_id=%s", (chat_id))
                result = mycursor.fetchall()
                for row in result:
                    find_area = row['area']
                    find_comment = row['comment']
                    find_maillist = row['maillist']
                    find_reg = row['registration']
                    if find_area == "mvu":
                        find_area = "МВУ"
                    elif find_area == "ru":
                        find_area = "РУ"
                    elif find_area == "du":
                        find_area = "ДУ"
                    elif find_area == "tu":
                        find_area = "ТУ"
                    elif find_area == "svu":
                        find_area = "СВУ"
                    elif find_area == "sku":
                        find_area = "СКУ"
                    elif find_area == "del":
                        find_area = "Заблокирован"
                    if find_maillist == 1:
                        find_maillist = "✅ Подписан"
                        action_maillist = "🛑 Отписаться"
                    else:
                        find_maillist = "🛑 Не подписан"
                        action_maillist = "✅ Подписаться"
                mydb.commit()
                mydb.close()
                set_sub = types.InlineKeyboardMarkup(row_width=1)
                button1 = types.InlineKeyboardButton("📤 Рассылка:  " + str(action_maillist), callback_data='call_sub_maillist')
                button2 = types.InlineKeyboardButton("⬅ Меню  ", callback_data='mainmenu')
                set_sub.add(button1, button2)
                bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id, text="О пользователе:\n🆔: " + str(chat_id) + "\n👤ФИО: " + find_comment + "\n🚈 Участок: " + find_area + "\n📤 Рассылка: " + str(find_maillist) + "\n🕐Дата регистрации: " + str(find_reg), parse_mode="Markdown", reply_markup=set_sub)
            if call.data == "call_sub_maillist":
                chat_id = call.message.chat.id
                firstname = call.from_user.first_name
                username = call.from_user.username
                mydb.ping(reconnect=True)  
                mycursor = mydb.cursor()  
                mycursor.execute("SELECT * FROM users WHERE chat_id=%s", (chat_id))
                result = mycursor.fetchall()
                for row in result:  
                    find_maillist = row["maillist"]
                    if find_maillist == 1:
                        mycursor.execute("UPDATE users SET maillist = '0' WHERE chat_id=%s", (chat_id))  
                        comment = "🛑 Уведомления отключены"
                        mydb.commit()  
                    else:
                        mycursor.execute("UPDATE users SET maillist = '1' WHERE chat_id=%s", (chat_id))  
                        comment = "✅ Уведомления включены"
                        mydb.commit()  

                bot.send_message(chat_id=chat_id, text='🐙 Рассылка - ' + comment)
                mydb.close()               
def areaopros(message):
    mydb.ping(reconnect=True)  
    mycursor = mydb.cursor()  
    mycursor.execute("SELECT * FROM users WHERE chat_id=%s", (chat_id))
    result = mycursor.fetchall()
    for row in result:
        valid_area = row['area']
    mydb.commit()
    mydb.close()
    wb = openpyxl.load_workbook(localfolder + 'data/uchastok/' + valid_area + '/opros/opros.xlsx')
    ws = wb['opros']
    tema = ws['L2'].value
    temaansw0 = ws['L3'].value
    temaansw1 = ws['L4'].value
    temaansw2 = ws['L5'].value
    if message.text.startswith('❌ стоп'):
        bot.send_message(message.chat.id, text='🐙 Отмена операции', reply_markup=types.ReplyKeyboardRemove())
    else:
        bot.send_message(chat_id=chat_id, text='💡 Пользователь получит от бота опрос в котором может проголосовать 1 раз. Статистика голосования у пользователей отображается под кнопкой голосования.\n')
        bot.send_message(chat_id=chat_id, text='📊 Указать данные для опроса: opros.xlxs')
        bot.send_message(chat_id=chat_id, text='-------\nТема опроса: ' + str(tema) + '\n1.' + str(temaansw0) + '\n2.' + str(temaansw1) + '\n3.' + str(temaansw2) + '\n-------')
        msg = bot.send_message(message.chat.id, text='🐙 Выбранный участок: *текущий*', parse_mode='Markdown', reply_markup=kball.stopandnext)
        bot.register_next_step_handler(msg, areaopros1)
def areaopros1(message):
    chat_id = message.chat.id
    input_area = message.text
    mydb.ping(
        reconnect=True) 
    mycursor = mydb.cursor()  
    mycursor.execute("SELECT * FROM users WHERE chat_id=%s", (chat_id))
    result = mycursor.fetchall()
    for row in result:
        valid_area = row['area']
    mydb.commit()
    mydb.close()
    import time
    import datetime
    dtn = datetime.datetime.now()
    wb = openpyxl.load_workbook(localfolder + 'data/uchastok/' + valid_area + '/opros/opros.xlsx') 
    ws = wb['opros']  
    tema = ws['L2'].value  
    temaansw0 = ws['L3'].value  
    temaansw1 = ws['L4'].value  
    temaansw2 = ws['L5'].value  
    if input_area.lower() == "❌ стоп":
        bot.send_message(message.chat.id, text='🐙 Отмена операци', reply_markup=types.ReplyKeyboardRemove())
    else:
        try:
            mycursor = mydb.cursor()  
            mycursor.execute("DELETE FROM opros WHERE area=%s", (valid_area))  
            mydb.commit()
        except:
            bot.send_message(message.chat.id, text='🐙 Рассылка начата!', reply_markup=types.ReplyKeyboardRemove())
            mydb.ping(reconnect=True) 
            mycursor = mydb.cursor() 
            maillist = '1'
            mycursor.execute("SELECT * FROM users WHERE maillist=%s AND area=%s", (maillist, valid_area))
            result = mycursor.fetchall()
            for row in result:
                chat_id = row['chat_id']
                comment = row['comment']
                try:
                    poll = bot.send_poll(chat_id=str(row["chat_id"]), question=str(tema), options=[str(temaansw0), str(temaansw1), str(temaansw2)], is_anonymous=False, reply_markup=kball.areaoprosinline)
                    time.sleep(0.05)  
                    add = "REPLACE INTO opros(chat_id, comment, folder, area, poll_id) VALUES('" + str(chat_id) + "', '" + str(comment) + "', '-', '" + str(valid_area) + "', '" + str(poll.id) + "');"
                    mycursor.execute(add)  
                    mydb.commit()
                except:
                    print("Ошибка при добавлении в БД opros, нет чата с пользователем")
        mydb.close()
        bot.send_message(message.chat.id, text='🐙 Рассылка завершена!', reply_markup=kball.areaopros)
        log.writelog(user_id=message.chat.id, username=message.from_user.first_name, text=" /send>ОПРОС: " + str(tema))
    if valid_area.lower() == "mvu":
        add_area = "МВУ"
    elif valid_area.lower() == "ru":
        add_area = "РУ"
    elif valid_area.lower() == "du":
        add_area = "ДУ"
    elif valid_area.lower() == "tu":
        add_area = "ТУ"
    elif valid_area.lower() == "svu":
        add_area = "СВУ"
    elif valid_area.lower() == "sku":
        add_area = "СКУ"

    opros = ws['A1'] = 'Дата проведения опроса:'
    opros = ws['B1'] = dtn.strftime("%d-%m-%Y %H:%M")
    opros = ws['A2'] = 'Участок:'
    opros = ws['B2'] = str(add_area)
    opros = ws['K2'] = 'Тема опроса:'
    opros = ws['K3'] = 'Варианты ответов:'
    wb.save(localfolder + 'data/uchastok/' + valid_area + '/opros/opros.xlsx')
    
@bot.poll_answer_handler(func=lambda call: True)
def hadle_poll(call):
     chat_id = call.user.id
     import datetime
     dtn = datetime.datetime.now()
     try:
        mycursor = mydb.cursor()  
        mydb.ping(reconnect=True)  
        mycursor.execute("SELECT area FROM opros WHERE chat_id=%s", (chat_id))
        result = mycursor.fetchone()
        area = str(result['area'])
        mydb.close()
        if area in ["vsem"]:
            wb = openpyxl.load_workbook(localfolder + 'data/uchastok/all/opros/opros.xlsx')
            ws = wb['opros']
            opros = [
                [dtn.strftime("%d-%m-%Y %H:%M"), str(call.user.first_name), str(call.user.id), str(call.option_ids[0])]]
            for info in opros:
                ws.append(info)
            ws.row_dimensions[1].height = 15
            ws.column_dimensions['A'].width = 24
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 14
            ws.column_dimensions['D'].width = 9
            ws.column_dimensions['K'].width = 18
            ws.column_dimensions['L'].width = 40
            ws["M3"] = '=COUNTIF(D6:D100, "0")'
            ws["M4"] = '=COUNTIF(D6:D100, "1")'
            ws["M5"] = '=COUNTIF(D6:D100, "2")'
            ws["M7"] = "=SUM(M3, M4, M5)"
            wb.save(localfolder + 'data/uchastok/all/opros/opros.xlsx')
        elif area in ["mvu", "ru", "du", "tu", "svu", "sku"]:
            wb = openpyxl.load_workbook(localfolder + 'data/uchastok/' + area + '/opros/opros.xlsx')
            ws = wb['opros']
            opros = [
                [dtn.strftime("%d-%m-%Y %H:%M"), str(call.user.first_name), str(call.user.id), str(call.option_ids[0])]]
            for info in opros:
                ws.append(info)
            ws.row_dimensions[1].height = 15
            ws.column_dimensions['A'].width = 24
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 14
            ws.column_dimensions['D'].width = 9
            ws.column_dimensions['K'].width = 18
            ws.column_dimensions['L'].width = 40
            ws["M3"] = '=COUNTIF(D6:D100, "0")'
            ws["M4"] = '=COUNTIF(D6:D100, "1")'
            ws["M5"] = '=COUNTIF(D6:D100, "2")'
            ws["M7"] = "=SUM(M3, M4, M5)"
            wb.save(localfolder + 'data/uchastok/' + area + '/opros/opros.xlsx')
        else:
            print("Опрос просрочился")
     except:
        print("Опрос просрочился")
        bot.send_message(call.user.id, text="🐙 Опрос просрочился!", reply_markup=types.ReplyKeyboardRemove())



context = ssl.SSLContext(ssl.PROTOCOL_TLSv1_2)
context.load_cert_chain(WEBHOOK_SSL_CERT, WEBHOOK_SSL_PRIV)

web.run_app(
    app,
    host=WEBHOOK_LISTEN,
    port=WEBHOOK_PORT,
    ssl_context=context,
)

